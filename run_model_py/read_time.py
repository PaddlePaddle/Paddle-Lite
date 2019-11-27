# coding=UTF-8
import sys
import os
import math
import numpy as np

import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy

def write_data(files, path):
    first = True
    for file_path in files:
        name = file_path.split('.')[0]
        print('name: ', name)
        data = read_txt(file_path)
        title = [[name, 'model', '1', '2', '4', ' ', '1', '2', '4'],]
        if first:
            write_excel_xls(path, 'sheet1', title)
            write_excel_xls_append(path, data, 0, 1)
            first = False
        else:
            write_excel_xls_append(path, title, 3, 0)
            write_excel_xls_append(path, data, 0, 1)
        
def read_txt(file_path):
    fp = open(file_path, 'r')
    if (not fp):
        print('open file failed', file_path)
    line = fp.readline()
    val_dict = {}
    while line:
        # data=model: ./model/merge21_ssd_shufflenet_quant-fluild, threads: 4, avg: 3.928300 ms, var: 0.267773
        data = line.strip().split(',')
        model = data[0].split(':')[1]
        model = model.split('/')[-1]
        print('model: ', model)
        val = {}
        if len(val_dict) == 0 or model not in val_dict:
            val = {}
        else:
            val = val_dict[model]
        temp = {}
        thread = data[1].split(':')[1].strip()
        temp['avg'] = data[2].split(':')[1].strip().split(' ')[0]
        # temp['min'] = data[3].split(':')[1].strip().split(' ')[0]
        temp['var'] = data[3].split(':')[1].strip()
        val[thread] = temp
        val_dict[model] = val
        line = fp.readline()
    
    print(val_dict)
    fp.close()
    data = []
    for key in val_dict.keys():
        val = val_dict[key]
        temp = []
        temp.append(key)
        temp.append(val['1']['avg'])
        temp.append(val['2']['avg'])
        temp.append(val['4']['avg'])
        temp.append('  ')
        temp.append(val['1']['var'])
        temp.append(val['2']['var'])
        temp.append(val['4']['var'])
        data.append(temp)
    print('----')
    print(data)
    return data

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook() #new excel book
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    
    workbook.save(path)

def write_excel_xls(path, sheet_name, value):
    index = len(value)
    workbook = xlwt.Workbook() #new excel book
    sheet = workbook.add_sheet(sheet_name)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    
    workbook.save(path)

def write_excel_xls_append(path, value, num=0, dtype=0):
    '''
    param num: rows
    param dtype: 0-title or 1-content
    '''
    index = len(value)
    workbook = xlrd.open_workbook(path) #open excel book
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    rows_old = worksheet.nrows
    new_workbook = copy(workbook)
    new_worksheet = new_workbook.get_sheet(0)
    
    rows_old = rows_old + num

    for i in range(0, index):
        for j in range(0, len(value[i])):
            if dtype:
                new_worksheet.write(i + rows_old, j+1, value[i][j])
            else:
                new_worksheet.write(i + rows_old, j, value[i][j])
    
    new_workbook.save(path)

if __name__ == '__main__':
    files = ['time_17c3cc34_v7.txt', 'time_17c3cc34_v8.txt', 'time_5380268d_v7.txt', 'time_5380268d_v8.txt', 'time_7f1446bd_v7.txt', 'time_7f1446bd_v8.txt']
    path = sys.argv[1]
    write_data(files, path)